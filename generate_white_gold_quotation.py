import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_white_and_gold_workbook(output_filename="DECOR8_INDIA_QUOTATION_WHITE_AND_GOLD.xlsx"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # ==========================================
    # --- LUXURY WHITE & GOLD COLOR PALETTE ---
    # ==========================================
    GOLD_PRIMARY   = "FFD4AF37"   # Metallic Luxury Gold (Headers, Banners, Grand Total)
    GOLD_DEEP      = "FFB8860B"   # Deep Warm Gold (Title, Brand Accents)
    GOLD_TBL_HDR   = "FFF5E8C2"   # Champagne Gold (Table Headers)
    GOLD_SUBTOTAL  = "FFEEDEA6"   # Soft Gold (Subtotals, Category Summaries)
    GOLD_DISCOUNT  = "FFFCE8E6"   # Soft Rose Tint or Warm Gold Accent for Discount
    GOLD_ZEBRA     = "FFFAF5E8"   # Subtle Warm Gold Tint (Alternating Rows)
    GOLD_META_LBL  = "FFF2E3BC"   # Gold Tint for Metadata & Bank Labels
    WHITE          = "FFFFFFFF"   # Pure Crisp White
    TEXT_DARK      = "FF1A1A1A"   # Crisp Jet Black
    TEXT_MUTED     = "FF555555"   # Subtitle & Note Text
    BORDER_GOLD    = "FFC9A232"   # Refined Gold Border
    BORDER_LIGHT   = "FFE2CE96"   # Subtle Inner Grid Border

    thin_gold = Side(border_style="thin", color=BORDER_GOLD)
    medium_gold = Side(border_style="medium", color=BORDER_GOLD)
    double_gold = Side(border_style="double", color=BORDER_GOLD)

    cell_border = Border(left=thin_gold, right=thin_gold, top=thin_gold, bottom=thin_gold)
    subtotal_border = Border(left=thin_gold, right=thin_gold, top=thin_gold, bottom=medium_gold)
    grand_total_border = Border(left=medium_gold, right=medium_gold, top=medium_gold, bottom=double_gold)

    # ==========================================
    # --- TYPOGRAPHY ---
    # ==========================================
    font_quotation_title = Font(name="Segoe UI", size=16, bold=True, color=GOLD_DEEP)
    font_company_name    = Font(name="Segoe UI", size=15, bold=True, color=TEXT_DARK)
    font_company_sub     = Font(name="Segoe UI", size=9, bold=False, color=TEXT_MUTED)
    font_gstin           = Font(name="Segoe UI", size=9.5, bold=True, color=TEXT_DARK)
    font_brand_rt        = Font(name="Georgia", size=16, bold=True, color=GOLD_DEEP)
    font_tagline_rt      = Font(name="Segoe UI", size=8.5, italic=True, color=TEXT_MUTED)

    font_banner          = Font(name="Segoe UI", size=10, bold=True, color=TEXT_DARK)
    font_sec_hdr         = Font(name="Segoe UI", size=10.5, bold=True, color=TEXT_DARK)
    font_tbl_hdr         = Font(name="Segoe UI", size=9.5, bold=True, color=TEXT_DARK)
    
    font_cell            = Font(name="Segoe UI", size=9, bold=False, color=TEXT_DARK)
    font_cell_bold       = Font(name="Segoe UI", size=9, bold=True, color=TEXT_DARK)
    font_cell_amount     = Font(name="Segoe UI", size=9.5, bold=True, color=TEXT_DARK)
    font_subtotal        = Font(name="Segoe UI", size=9.5, bold=True, color=TEXT_DARK)
    font_grand_total     = Font(name="Segoe UI", size=11, bold=True, color=TEXT_DARK)
    
    font_terms_hdr       = Font(name="Segoe UI", size=9.5, bold=True, color=TEXT_DARK)
    font_terms_text      = Font(name="Segoe UI", size=8.5, bold=False, color="FF222222")
    font_footer          = Font(name="Segoe UI", size=9.5, bold=True, color=TEXT_DARK)

    # ==========================================
    # --- FILLS ---
    # ==========================================
    fill_gold_hdr    = PatternFill(start_color=GOLD_PRIMARY, end_color=GOLD_PRIMARY, fill_type="solid")
    fill_tbl_hdr     = PatternFill(start_color=GOLD_TBL_HDR, end_color=GOLD_TBL_HDR, fill_type="solid")
    fill_subtotal    = PatternFill(start_color=GOLD_SUBTOTAL, end_color=GOLD_SUBTOTAL, fill_type="solid")
    fill_grand_total = PatternFill(start_color=GOLD_PRIMARY, end_color=GOLD_PRIMARY, fill_type="solid")
    fill_zebra_light = PatternFill(start_color=GOLD_ZEBRA, end_color=GOLD_ZEBRA, fill_type="solid")
    fill_white       = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    fill_meta_lbl    = PatternFill(start_color=GOLD_META_LBL, end_color=GOLD_META_LBL, fill_type="solid")

    # ==========================================
    # --- ALIGNMENTS ---
    # ==========================================
    align_center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left      = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    align_right     = Alignment(horizontal="right", vertical="center", indent=1, wrap_text=True)
    align_left_top  = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)

    def style_and_merge(ws, start_col, start_row, end_col, end_row, value=None, font=None, fill=None, border=cell_border, alignment=None, num_format=None):
        for r_idx in range(start_row, end_row + 1):
            for c_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if font is not None: cell.font = font
                if fill is not None: cell.fill = fill
                if border is not None: cell.border = border
                if alignment is not None: cell.alignment = alignment
                if num_format is not None: cell.number_format = num_format
        
        top_left = ws.cell(row=start_row, column=start_col)
        if value is not None:
            top_left.value = value
            
        if start_col != end_col or start_row != end_row:
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

    def build_quotation_sheet(sheet_title, client_name="MR.UDAY", quote_no="D8202602105", quote_date="16-Feb-2026", rep="Mr.Satish", contact="+91", address="Bengaluru", email=""):
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True

        column_widths = {
            'A': 8,     # Sl No
            'B': 22,    # Item Name / Particulars
            'C': 34,    # Sub Item Name / Item Details
            'D': 14,    # Total SQFT
            'E': 16,    # Rate / Unit Price
            'F': 18,    # Amount (₹)
            'G': 40     # Remarks
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        r = 1

        # ----------------------------------------------------
        # 1. Top Header Title
        # ----------------------------------------------------
        ws.row_dimensions[r].height = 28
        style_and_merge(ws, 1, r, 7, r, value="QUOTATION", font=font_quotation_title, fill=fill_white, border=cell_border, alignment=align_center)
        r += 1

        # ----------------------------------------------------
        # 2. Company Info (Left) & Brand Identity (Right)
        # ----------------------------------------------------
        ws.row_dimensions[r].height = 20
        ws.row_dimensions[r+1].height = 16
        ws.row_dimensions[r+2].height = 16
        ws.row_dimensions[r+3].height = 16
        ws.row_dimensions[r+4].height = 18

        ws[f"A{r}"] = "DECOR8 INDIA"
        ws[f"A{r}"].font = font_company_name
        ws[f"A{r}"].alignment = align_left

        ws[f"A{r+1}"] = "#14, Sy NO 36/1 Vasanth Vallabhnagar, Vasanthapura"
        ws[f"A{r+1}"].font = font_company_sub
        ws[f"A{r+1}"].alignment = align_left

        ws[f"A{r+2}"] = "Uttarahalli Hobli ,Bengaluru - 560061"
        ws[f"A{r+2}"].font = font_company_sub
        ws[f"A{r+2}"].alignment = align_left

        ws[f"A{r+3}"] = "8884131414 ,9380523743"
        ws[f"A{r+3}"].font = font_company_sub
        ws[f"A{r+3}"].alignment = align_left

        ws[f"A{r+4}"] = "GSTIN: 29CHPPB0948C2ZD"
        ws[f"A{r+4}"].font = font_gstin
        ws[f"A{r+4}"].alignment = align_left

        # Right: Brand Identity Block
        style_and_merge(ws, 5, r, 7, r+1, value="DECOR8 INDIA", font=font_brand_rt, alignment=Alignment(horizontal="right", vertical="center"), border=None)
        style_and_merge(ws, 5, r+2, 7, r+3, value="Affordable Luxury Interiors\nwww.decor8india.com", font=font_tagline_rt, alignment=Alignment(horizontal="right", vertical="center", wrap_text=True), border=None)

        r += 5

        # Gold Divider Line
        ws.row_dimensions[r].height = 4
        style_and_merge(ws, 1, r, 7, r, fill=fill_gold_hdr, border=None)
        r += 1

        # ----------------------------------------------------
        # 3. Client & Quotation Metadata
        # ----------------------------------------------------
        ws.row_dimensions[r].height = 24
        lbl_q_for = f"QUOTATION For - {client_name}"
        style_and_merge(ws, 1, r, 4, r, value=lbl_q_for, font=font_banner, fill=fill_gold_hdr, border=cell_border, alignment=align_left)
        style_and_merge(ws, 5, r, 7, r, value=f"Company Representative: {rep}", font=font_banner, fill=fill_gold_hdr, border=cell_border, alignment=align_left)
        r += 1

        meta_rows = [
            ("Quotations No:", quote_no, "Quotations Date:", quote_date),
            ("Contact No:", contact, "Address:", address),
            ("Email ID :", email, "", "")
        ]

        for lbl1, val1, lbl2, val2 in meta_rows:
            ws.row_dimensions[r].height = 20
            # Left Label & Value
            style_and_merge(ws, 1, r, 1, r, value=lbl1, font=font_cell_bold, fill=fill_meta_lbl, border=cell_border, alignment=align_left)
            style_and_merge(ws, 2, r, 4, r, value=val1, font=font_cell, fill=fill_white, border=cell_border, alignment=align_left)

            # Right Label & Value
            style_and_merge(ws, 5, r, 5, r, value=lbl2, font=font_cell_bold, fill=fill_meta_lbl if lbl2 else fill_white, border=cell_border, alignment=align_left)
            style_and_merge(ws, 6, r, 7, r, value=val2, font=font_cell, fill=fill_white, border=cell_border, alignment=align_left)
            r += 1

        # Spacing
        ws.row_dimensions[r].height = 6
        r += 1

        # ----------------------------------------------------
        # 4. Section A: Wood work & Modular Finish CENTURY PLY
        # ----------------------------------------------------
        ws.row_dimensions[r].height = 24
        style_and_merge(ws, 1, r, 7, r, value="Section A: Wood work & Modular Finish CENTURY PLY", font=font_sec_hdr, fill=fill_gold_hdr, border=cell_border, alignment=align_left)
        r += 1

        # Section A Headers
        ws.row_dimensions[r].height = 22
        headers_a = ["Sl\nNo", "Item Name", "Sub Item Name", "Total\nSQFT", "Rate", "Amount", "Remarks"]
        for c_idx, h_text in enumerate(headers_a, start=1):
            style_and_merge(ws, c_idx, r, c_idx, r, value=h_text, font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_center if c_idx in [1, 4, 5, 6] else align_left)
        r += 1

        sec_a_start_row = r

        # Exact items from Screenshot 1 (MR.UDAY)
        # Groups: (SlNo, ItemName, [(SubItem, SQFT, Rate, Remarks), ...])
        sec_a_groups = [
            (1, "Modular Kitchen", [
                ("Base Cabinet", 43, 1600, "BWP Ply Carcases & HDHMR Shutters with Premium Acrylic finish"),
                ("Wall Cabinet", 26, 1600, ""),
                ("Tall Unit", 9, 1100, ""),
                ("Loft", 43, 1100, "")
            ]),
            (2, "MBR", [
                ("Swing wardrobe", 56, 1600, "MR Ply carcase with 18mm HDHMR shutters & Acrylic finish"),
                ("Loft", 18, 1100, ""),
                ("Study with dressing", 56, 1100, "As per old design")
            ]),
            (3, "KBR", [
                ("Swing Wardrobe", 49, 1600, "MR Ply carcase with 18mm HDHMR shutters & Acrylic finish"),
                ("Loft", 14, 1100, ""),
                ("StudyTable", 42, 1200, "Acrylic finish")
            ]),
            (4, "Living & Dining Area", [
                ("TV Unit", 96, 1950, "Acrylic and Louvers combination"),
                ("Crockery", 21, 1900, "Profile shutters with Acrylic carcase"),
                ("Pooja Unit", 21, 1950, "CNC design with Acrylic finish"),
                ("Wall Panelling", 96, 550, "As per 3D")
            ]),
            (5, "GBR", [
                ("Wardrobe", 35, 1600, "MR Ply carcase & 18MM HDHMR shutters with Acrylic finish"),
                ("Loft", 10, 1100, ""),
                ("Study", 42, 1200, "")
            ]),
            (6, "Foyer Area", [
                ("Shoe Unit", 14, 1400, "Seating cushion included"),
                ("Wall Highlight with charcoal panels", 50, 650, "Entrance wall"),
                ("Bronze mirror", 20, 950, "")
            ])
        ]

        for sl_no, item_name, subitems in sec_a_groups:
            group_start = r
            for idx, (subitem, sqft, rate, rem) in enumerate(subitems):
                ws.row_dimensions[r].height = 20
                ws[f"A{r}"] = sl_no if idx == 0 else ""
                ws[f"B{r}"] = item_name if idx == 0 else ""
                ws[f"C{r}"] = subitem
                ws[f"D{r}"] = sqft
                ws[f"E{r}"] = rate
                ws[f"F{r}"] = f"=D{r}*E{r}"
                ws[f"G{r}"] = rem
                r += 1
            group_end = r - 1
            if group_end > group_start:
                ws.merge_cells(f"A{group_start}:A{group_end}")
                ws.merge_cells(f"B{group_start}:B{group_end}")

        sec_a_end_row = r - 1

        for row_idx in range(sec_a_start_row, sec_a_end_row + 1):
            ws[f"A{row_idx}"].alignment = align_center
            ws[f"B{row_idx}"].alignment = align_center
            ws[f"C{row_idx}"].alignment = align_left
            ws[f"D{row_idx}"].alignment = align_center
            ws[f"E{row_idx}"].alignment = align_right
            ws[f"F{row_idx}"].alignment = align_right
            ws[f"G{row_idx}"].alignment = align_left

            ws[f"A{row_idx}"].font = font_cell
            ws[f"B{row_idx}"].font = font_cell_bold
            ws[f"C{row_idx}"].font = font_cell
            ws[f"D{row_idx}"].font = font_cell
            ws[f"E{row_idx}"].font = font_cell
            ws[f"F{row_idx}"].font = font_cell_amount
            ws[f"G{row_idx}"].font = font_cell

            is_even = (row_idx % 2 == 0)
            for col_idx in range(1, 8):
                col_letter = get_column_letter(col_idx)
                ws[f"{col_letter}{row_idx}"].border = cell_border
                ws[f"{col_letter}{row_idx}"].fill = fill_zebra_light if is_even else fill_white

            ws[f"D{row_idx}"].number_format = "#,##0"
            ws[f"E{row_idx}"].number_format = "₹ #,##0.00"
            ws[f"F{row_idx}"].number_format = "₹ #,##0.00"

        # Sub Total Section A (excluding GST)
        ws.row_dimensions[r].height = 22
        style_and_merge(ws, 1, r, 5, r, value="Sub Total excluding GST", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
        style_and_merge(ws, 6, r, 6, r, value=f"=SUM(F{sec_a_start_row}:F{sec_a_end_row})", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right, num_format="₹ #,##0.00")
        style_and_merge(ws, 7, r, 7, r, fill=fill_subtotal, border=subtotal_border)
        subtotal_a_raw_row = r
        r += 1

        # Discount Row (Exact from Screenshot 1: ₹ 1,00,000)
        ws.row_dimensions[r].height = 22
        style_and_merge(ws, 1, r, 5, r, value="Discount", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
        style_and_merge(ws, 6, r, 6, r, value=100000, font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right, num_format="₹ #,##0.00")
        style_and_merge(ws, 7, r, 7, r, fill=fill_subtotal, border=subtotal_border)
        discount_a_row = r
        r += 1

        # Final Sub Total Section A (A - Discount)
        ws.row_dimensions[r].height = 22
        style_and_merge(ws, 1, r, 5, r, value="Sub Total excluding GST (A)", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
        style_and_merge(ws, 6, r, 6, r, value=f"=F{subtotal_a_raw_row}-F{discount_a_row}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right, num_format="₹ #,##0.00")
        style_and_merge(ws, 7, r, 7, r, fill=fill_subtotal, border=subtotal_border)
        subtotal_a_row = r
        r += 1

        # Spacing
        ws.row_dimensions[r].height = 6
        r += 1

        # ----------------------------------------------------
        # 5. Section B: Appliances & Accessories
        # ----------------------------------------------------
        ws.row_dimensions[r].height = 24
        style_and_merge(ws, 1, r, 7, r, value="Section B: Appliances & Accessories", font=font_sec_hdr, fill=fill_gold_hdr, border=cell_border, alignment=align_left)
        r += 1

        ws.row_dimensions[r].height = 22
        style_and_merge(ws, 1, r, 1, r, value="Sl\nNo", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_center)
        style_and_merge(ws, 2, r, 2, r, value="Particulars", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_left)
        style_and_merge(ws, 3, r, 4, r, value="Item Details", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_left)
        style_and_merge(ws, 5, r, 5, r, value="Price", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_center)
        style_and_merge(ws, 6, r, 6, r, value="Amount", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_center)
        style_and_merge(ws, 7, r, 7, r, value="Remarks", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_left)
        r += 1

        sec_b_start_row = r

        # Kitchen Accessories (Sl No 1)
        kitchen_items = [
            ("Hob and Chimney (Elica /Kutchina)", "-", ""),
            ("Sink", "-", ""),
            ("Inbuilt Oven and Microwave", "-", "As per model selection price may vary"),
            ("Copper pipe for Gas Cylender connection", "-", ""),
            ("1x cutlery with tandem drawer", "", ""),
            ("1x plain ss basket drawer", "", ""),
            ("Rolling shutter", "", ""),
            ("1x thali basket , Magic corner , Wicker basket", "", "")
        ]

        kitchen_start = r
        for idx, (item_det, price_val, rem) in enumerate(kitchen_items):
            ws.row_dimensions[r].height = 20
            ws[f"A{r}"] = 1 if idx == 0 else ""
            ws[f"B{r}"] = "Kitchen\nHardwares(including\nsoft hinges )" if idx == 0 else ""
            ws.merge_cells(f"C{r}:D{r}")
            ws[f"C{r}"] = item_det
            ws[f"E{r}"] = price_val if idx < 4 else ""
            ws[f"F{r}"] = price_val if idx < 4 else (47850 if idx == 4 else "")
            ws[f"G{r}"] = rem
            r += 1
        kitchen_end = r - 1

        ws.merge_cells(f"A{kitchen_start}:A{kitchen_end}")
        ws.merge_cells(f"B{kitchen_start}:B{kitchen_end}")
        # Merge the 47,850 amount across the 4 basket rows
        ws.merge_cells(f"F{kitchen_start+4}:F{kitchen_end}")
        ws.merge_cells(f"E{kitchen_start+4}:E{kitchen_end}")

        # Wardrobes Accessories (Sl No 2)
        wardrobe_start = r
        wardrobe_items = [
            ("Hardware ( 3 wardrobes)", 30300, "handles, hinges and required hardwares added"),
            ("", "-", ""),
            ("", "", ""),
            ("", "", "")
        ]

        for idx, (item_det, price_val, rem) in enumerate(wardrobe_items):
            ws.row_dimensions[r].height = 20
            ws[f"A{r}"] = 2 if idx == 0 else ""
            ws[f"B{r}"] = "Wardrobes" if idx == 0 else ""
            ws.merge_cells(f"C{r}:D{r}")
            ws[f"C{r}"] = item_det
            ws[f"E{r}"] = price_val
            ws[f"F{r}"] = price_val
            ws[f"G{r}"] = rem
            r += 1
        wardrobe_end = r - 1

        ws.merge_cells(f"A{wardrobe_start}:A{wardrobe_end}")
        ws.merge_cells(f"B{wardrobe_start}:B{wardrobe_end}")

        sec_b_end_row = r - 1

        for row_idx in range(sec_b_start_row, sec_b_end_row + 1):
            ws[f"A{row_idx}"].alignment = align_center
            ws[f"B{row_idx}"].alignment = align_center
            ws[f"C{row_idx}"].alignment = align_left
            ws[f"E{row_idx}"].alignment = align_center if str(ws[f"E{row_idx}"].value).strip() in ["-", ""] else align_right
            ws[f"F{row_idx}"].alignment = align_center if str(ws[f"F{row_idx}"].value).strip() in ["-", ""] else align_right
            ws[f"G{row_idx}"].alignment = align_left

            for c in range(1, 8):
                c_l = get_column_letter(c)
                ws[f"{c_l}{row_idx}"].font = font_cell
                ws[f"{c_l}{row_idx}"].border = cell_border
                ws[f"{c_l}{row_idx}"].fill = fill_zebra_light if row_idx % 2 == 0 else fill_white

            if isinstance(ws[f"E{row_idx}"].value, (int, float)):
                ws[f"E{row_idx}"].number_format = "₹ #,##0.00"
                ws[f"E{row_idx}"].font = font_cell_bold
            if isinstance(ws[f"F{row_idx}"].value, (int, float)):
                ws[f"F{row_idx}"].number_format = "₹ #,##0.00"
                ws[f"F{row_idx}"].font = font_cell_amount

        # Sub Total Section B
        ws.row_dimensions[r].height = 22
        style_and_merge(ws, 1, r, 5, r, value="Sub Total excluding GST (B)", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
        style_and_merge(ws, 6, r, 6, r, value=f"=SUM(F{kitchen_start+4}, F{wardrobe_start})", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right, num_format="₹ #,##0.00")
        style_and_merge(ws, 7, r, 7, r, fill=fill_subtotal, border=subtotal_border)
        subtotal_b_row = r
        r += 1

        # Spacing
        ws.row_dimensions[r].height = 6
        r += 1

        # ----------------------------------------------------
        # 6. Section C: Miscellaneous & Loose Furniture
        # ----------------------------------------------------
        ws.row_dimensions[r].height = 24
        style_and_merge(ws, 1, r, 7, r, value="Section C: Miscellaneous & Loose Furniture", font=font_sec_hdr, fill=fill_gold_hdr, border=cell_border, alignment=align_left)
        r += 1

        ws.row_dimensions[r].height = 22
        style_and_merge(ws, 1, r, 1, r, value="Sl\nNo", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_center)
        style_and_merge(ws, 2, r, 4, r, value="Particulars", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_left)
        style_and_merge(ws, 5, r, 5, r, value="Price", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_center)
        style_and_merge(ws, 6, r, 6, r, value="Amount", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_center)
        style_and_merge(ws, 7, r, 7, r, value="Remarks", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_left)
        r += 1

        misc_items = [
            (1, "KBR Cot background Design", 45500, "70 sqft total x 650 Rs"),
            (2, "Washroom Glass partition", 34000, ""),
            (3, "King size bed Hydrolic with Headboard and side table", 58000, "Acrylic finish"),
            (4, "Queen size bed Hydrolic with Headboard and side table with cushion", 56000, ""),
            (5, "Wallpapers", 15000, ""),
            (6, "Painting (Royal Aspira ) 1 coat primer 2coat Asian paint Aspira", 48850, ""),
            (7, "Quartz/Granite stone including cutting and installation", 7500, ""),
            (8, "MBR cot background (louvers and arch )", 55900, "86 sqft x 650"),
            (9, "False ceiling , Wall Mouldings as per design", 62000, ""),
            ("", "Wall Mouldings as per design", 16500, "including electrical shifting"),
            (10, "Electrical work ( including pop wiring switches etc )", 60000, "Pop Lights, Profile lights installation with materials"),
            (11, "GBR louvers", 29250, "65 sqft x 450")
        ]

        sec_c_start_row = r
        for idx_sl, particulars, amt, rem in misc_items:
            ws.row_dimensions[r].height = 20
            ws[f"A{r}"] = idx_sl
            ws.merge_cells(f"B{r}:D{r}")
            ws[f"B{r}"] = particulars
            ws[f"E{r}"] = amt
            ws[f"F{r}"] = amt
            ws[f"G{r}"] = rem
            r += 1
        sec_c_end_row = r - 1

        for row_idx in range(sec_c_start_row, sec_c_end_row + 1):
            ws[f"A{row_idx}"].alignment = align_center
            ws[f"B{row_idx}"].alignment = align_left
            ws[f"E{row_idx}"].alignment = align_right
            ws[f"F{row_idx}"].alignment = align_right
            ws[f"G{row_idx}"].alignment = align_left

            for c in range(1, 8):
                c_l = get_column_letter(c)
                ws[f"{c_l}{row_idx}"].font = font_cell
                ws[f"{c_l}{row_idx}"].border = cell_border
                ws[f"{c_l}{row_idx}"].fill = fill_zebra_light if row_idx % 2 == 0 else fill_white

            ws[f"E{row_idx}"].number_format = "₹ #,##0.00"
            ws[f"F{row_idx}"].number_format = "₹ #,##0.00"
            ws[f"F{row_idx}"].font = font_cell_amount

        # Sub Total Section C
        ws.row_dimensions[r].height = 22
        style_and_merge(ws, 1, r, 5, r, value="Sub Total excluding GST (C)", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
        style_and_merge(ws, 6, r, 6, r, value=f"=SUM(F{sec_c_start_row}:F{sec_c_end_row})", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right, num_format="₹ #,##0.00")
        style_and_merge(ws, 7, r, 7, r, fill=fill_subtotal, border=subtotal_border)
        subtotal_c_row = r
        r += 1

        # Spacing
        ws.row_dimensions[r].height = 6
        r += 1

        # ----------------------------------------------------
        # 7. Total Project Cost Summary
        # ----------------------------------------------------
        ws.row_dimensions[r].height = 24
        style_and_merge(ws, 1, r, 5, r, value="Total Project Cost (A+B+C) excluding GST", font=font_subtotal, fill=fill_subtotal, border=cell_border, alignment=align_right)
        style_and_merge(ws, 6, r, 6, r, value=f"=F{subtotal_a_row}+F{subtotal_b_row}+F{subtotal_c_row}", font=Font(name="Segoe UI", size=10.5, bold=True, color=TEXT_DARK), fill=fill_subtotal, border=cell_border, alignment=align_right, num_format="₹ #,##0.00")
        style_and_merge(ws, 7, r, 7, r, fill=fill_subtotal, border=cell_border)
        total_excl_gst_row = r
        r += 1

        # CGST 9%
        ws.row_dimensions[r].height = 20
        style_and_merge(ws, 1, r, 5, r, value="CGST @ 9%", font=font_cell_bold, fill=fill_white, border=cell_border, alignment=align_right)
        style_and_merge(ws, 6, r, 6, r, value=f"=ROUND(F{total_excl_gst_row}*0.09, 0)", font=font_cell_bold, fill=fill_white, border=cell_border, alignment=align_right, num_format="₹ #,##0.00")
        style_and_merge(ws, 7, r, 7, r, fill=fill_white, border=cell_border)
        cgst_row = r
        r += 1

        # SGST 9%
        ws.row_dimensions[r].height = 20
        style_and_merge(ws, 1, r, 5, r, value="SGST @ 9%", font=font_cell_bold, fill=fill_white, border=cell_border, alignment=align_right)
        style_and_merge(ws, 6, r, 6, r, value=f"=ROUND(F{total_excl_gst_row}*0.09, 0)", font=font_cell_bold, fill=fill_white, border=cell_border, alignment=align_right, num_format="₹ #,##0.00")
        style_and_merge(ws, 7, r, 7, r, fill=fill_white, border=cell_border)
        sgst_row = r
        r += 1

        # Total Project Cost including GST
        ws.row_dimensions[r].height = 26
        style_and_merge(ws, 1, r, 5, r, value="Total Project Cost (A+B+C) including GST", font=font_grand_total, fill=fill_grand_total, border=grand_total_border, alignment=align_right)
        style_and_merge(ws, 6, r, 6, r, value=f"=F{total_excl_gst_row}+F{cgst_row}+F{sgst_row}", font=Font(name="Segoe UI", size=12, bold=True, color=TEXT_DARK), fill=fill_grand_total, border=grand_total_border, alignment=align_right, num_format="₹ #,##0.00")
        style_and_merge(ws, 7, r, 7, r, fill=fill_grand_total, border=grand_total_border)
        grand_total_row = r
        r += 1

        # Spacing
        ws.row_dimensions[r].height = 6
        r += 1

        # ----------------------------------------------------
        # 8. Payment Schedule
        # ----------------------------------------------------
        ws.row_dimensions[r].height = 24
        style_and_merge(ws, 1, r, 7, r, value="Payment Schedule", font=font_sec_hdr, fill=fill_gold_hdr, border=cell_border, alignment=align_left)
        r += 1

        ws.row_dimensions[r].height = 22
        style_and_merge(ws, 1, r, 1, r, value="Sl\nNo", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_center)
        style_and_merge(ws, 2, r, 4, r, value="Payment Timelines", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_left)
        style_and_merge(ws, 5, r, 5, r, value="Percentage", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_center)
        style_and_merge(ws, 6, r, 7, r, value="Amount", font=font_tbl_hdr, fill=fill_tbl_hdr, border=cell_border, alignment=align_center)
        r += 1

        payment_milestones = [
            ("At the time of project/order confirmation", 0.15),
            ("At the time of starting the production/execution at factory", 0.50),
            ("After completion of production/execution at the factory or before send to the site for the installation", 0.35)
        ]

        for idx, (timeline, pct) in enumerate(payment_milestones, start=1):
            ws.row_dimensions[r].height = 24
            style_and_merge(ws, 1, r, 1, r, value=idx, font=font_cell, fill=fill_zebra_light if idx % 2 == 0 else fill_white, border=cell_border, alignment=align_center)
            style_and_merge(ws, 2, r, 4, r, value=timeline, font=font_cell, fill=fill_zebra_light if idx % 2 == 0 else fill_white, border=cell_border, alignment=align_left)
            style_and_merge(ws, 5, r, 5, r, value=pct, font=font_cell_bold, fill=fill_zebra_light if idx % 2 == 0 else fill_white, border=cell_border, alignment=align_center, num_format="0%")
            style_and_merge(ws, 6, r, 7, r, value=f"=ROUND(F{grand_total_row}*E{r}, 0)", font=font_cell_amount, fill=fill_zebra_light if idx % 2 == 0 else fill_white, border=cell_border, alignment=align_right, num_format="₹ #,##0.00")
            r += 1

        # Spacing
        ws.row_dimensions[r].height = 6
        r += 1

        # ----------------------------------------------------
        # 9. Company Account Information
        # ----------------------------------------------------
        ws.row_dimensions[r].height = 24
        style_and_merge(ws, 1, r, 7, r, value="Company Account information", font=font_sec_hdr, fill=fill_gold_hdr, border=cell_border, alignment=align_left)
        r += 1

        ws.row_dimensions[r].height = 18
        style_and_merge(ws, 1, r, 7, r, value="Please issue A/c Payee cheque in the name of \"DECOR8 INDIA\"", font=Font(name="Segoe UI", size=9, bold=True, color=TEXT_DARK), fill=fill_zebra_light, border=cell_border, alignment=align_center)
        r += 1

        ws.row_dimensions[r].height = 18
        style_and_merge(ws, 1, r, 7, r, value="You can also make online transactions for the below mentioned account details:", font=Font(name="Segoe UI", size=8.5, italic=True, color=TEXT_MUTED), fill=fill_zebra_light, border=cell_border, alignment=align_center)
        r += 1

        bank_details = [
            ("Bank Name:", "IDFC FIRST BANK"),
            ("Branch Name:", "BTM Layout"),
            ("A/C Name:", "DECOR8 INDIA"),
            ("A/C Number:", "10075641863"),
            ("A/C Type:", "Current"),
            ("IFSC Code:", "IDFB0080182")
        ]

        for lbl, val in bank_details:
            ws.row_dimensions[r].height = 20
            style_and_merge(ws, 1, r, 2, r, value=lbl, font=font_cell_bold, fill=fill_meta_lbl, border=cell_border, alignment=align_left)
            style_and_merge(ws, 3, r, 5, r, value=val, font=font_cell_bold if "DECOR8" in val or "IDFC" in val else font_cell, fill=fill_white, border=cell_border, alignment=align_left)
            style_and_merge(ws, 6, r, 7, r, fill=fill_white, border=cell_border)
            r += 1

        # Spacing
        ws.row_dimensions[r].height = 6
        r += 1

        # ----------------------------------------------------
        # 10. Terms & Conditions
        # ----------------------------------------------------
        ws.row_dimensions[r].height = 24
        style_and_merge(ws, 1, r, 7, r, value="* Terms & Conditions", font=font_terms_hdr, fill=fill_gold_hdr, border=cell_border, alignment=align_center)
        r += 1

        terms_list = [
            "Price is valid for 30 days only after the 1st estimation is offered. Possibility of price revision may happen due to vendor supplier reasons.",
            "DECOR8 will cover warranty period upto 15 years for any wood work related products on manufacturing defects (Plywood bendness not included). DECOR8 will not cover warranty for any intentional wear and tear or damages from the customer side. All other products from different vendors like Hettich, KAFF, EBCO, Havells etc will be covered by respective vendors /company as per their company norms and DECOR8 will support for the same.",
            "Payment terms as per payment slab should be followed from the time of booking by the client for the smooth functioning of work and output. After any revised quote due to any addition of products/changes, if price adds up, client needs to pay the difference amount as on the slab installment during that phase. This is due to the production related requirement for procuring materials, etc. For any reductions on amount due changes, the amount will be adjusted in the final billing stage.",
            "Designs will not start until complete booking amount is credited to the company followed by the payment slabs for production & installation after progress.",
            "Any delay in terms of payment as per above payment schedule then there might be a delay in the execution for which DECOR8 wont be responsible.",
            "Price may be revised if there is any hike in the raw material announce officially by the supplier if the project is still in designing stage and is yet to enter production stage. No revision of price for projects which are under production.",
            "Once an order is confirmed and customer wants to cancel the order then NO amount will be refunded.",
            "No hidden charges for 3D designs (only 03 corrections allowed)",
            "Any corrections beyond 03 times, INR 3000 per each rendering will be charged to the client.",
            "Once the designs are approved and execution has started, no changes will be accepted in terms of structure or designs."
        ]

        for idx, term in enumerate(terms_list, start=1):
            line_count = (len(term) // 110) + 1
            ws.row_dimensions[r].height = max(20, line_count * 16)
            style_and_merge(ws, 1, r, 7, r, value=f"* {term}", font=font_terms_text, fill=fill_white if idx % 2 != 0 else fill_zebra_light, border=cell_border, alignment=align_left_top)
            r += 1

        # Spacing
        ws.row_dimensions[r].height = 6
        r += 1

        # ----------------------------------------------------
        # 11. Footer Banner
        # ----------------------------------------------------
        ws.row_dimensions[r].height = 24
        style_and_merge(ws, 1, r, 7, r, value="www.decor8india.com", font=font_footer, fill=fill_gold_hdr, border=cell_border, alignment=align_center)
        r += 1

    # Create sheets
    build_quotation_sheet("Quotation - MR.UDAY", client_name="MR.UDAY", quote_no="D8202602105", quote_date="16-Feb-2026", rep="Mr.Satish", contact="+91", address="Bengaluru")
    build_quotation_sheet("Decor8 Luxury White & Gold", client_name="MR.UDAY", quote_no="D8202602105", quote_date="16-Feb-2026", rep="Mr.Satish", contact="+91", address="Bengaluru")

    # Save to requested destination and public folder
    try:
        wb.save(output_filename)
        print(f"Successfully saved White & Gold workbook: {output_filename}")
    except PermissionError:
        alt = output_filename.replace(".xlsx", "_White_Gold.xlsx")
        wb.save(alt)
        print(f"Locked file. Saved to alternate: {alt}")

    pub_path = os.path.join("public", os.path.basename(output_filename))
    try:
        wb.save(pub_path)
        print(f"Saved public copy: {pub_path}")
    except Exception as e:
        print("Public save error:", e)

    # Also save to luxury gold file name
    try:
        wb.save("Decor8India_Quotation_Luxury_Gold.xlsx")
        print("Saved Decor8India_Quotation_Luxury_Gold.xlsx")
    except Exception as e:
        print("Error saving luxury gold:", e)

if __name__ == "__main__":
    create_white_and_gold_workbook("DECOR8_INDIA_QUOTATION_WHITE_AND_GOLD.xlsx")
