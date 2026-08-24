import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_perfect_decor8_quotation(output_path="DECOR8_INDIA_QUOTATION_MR_GOVINDA_D1304.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # ==========================================
    # --- COLOR PALETTE (DECOR8 INDIA THEME) ---
    # ==========================================
    GOLD_PRIMARY = "D4AF37"     # Rich Luxury Gold
    GOLD_DARK = "997A15"        # Deep Gold
    GOLD_LIGHT = "FDFBF4"       # Soft Champagne tint
    GOLD_ACCENT = "EFE1BA"      # Table header fill
    GOLD_TOTAL = "E4D29F"       # Subtotal fill
    SLATE_DARK = "111317"       # Deep Obsidian
    SLATE_NAVY = "1C2028"       # Charcoal Navy
    WHITE = "FFFFFF"
    BORDER_COLOR = "C29B27"     # Gold border
    BORDER_SUBTLE = "DFC98A"    # Light Gold border
    TEXT_DARK = "111111"
    TEXT_MUTED = "444444"

    # ==========================================
    # --- BORDERS ---
    # ==========================================
    thin_gold = Side(border_style="thin", color=BORDER_COLOR)
    medium_gold = Side(border_style="medium", color=BORDER_COLOR)
    double_gold = Side(border_style="double", color=BORDER_COLOR)

    cell_border = Border(left=thin_gold, right=thin_gold, top=thin_gold, bottom=thin_gold)
    subtotal_border = Border(left=thin_gold, right=thin_gold, top=thin_gold, bottom=medium_gold)
    grand_total_border = Border(left=medium_gold, right=medium_gold, top=medium_gold, bottom=double_gold)

    # ==========================================
    # --- FONTS ---
    # ==========================================
    font_main_title = Font(name="Segoe UI", size=15, bold=True, color=WHITE)
    font_company = Font(name="Segoe UI", size=14, bold=True, color=SLATE_DARK)
    font_sub_info = Font(name="Segoe UI", size=9, bold=False, color=TEXT_MUTED)
    font_gstin = Font(name="Segoe UI", size=9.5, bold=True, color=SLATE_DARK)
    font_brand_rt = Font(name="Georgia", size=16, bold=True, color=GOLD_DARK)
    font_tagline_rt = Font(name="Segoe UI", size=8.5, italic=True, color=TEXT_MUTED)

    font_banner = Font(name="Segoe UI", size=10, bold=True, color=SLATE_DARK)
    font_sec_hdr = Font(name="Segoe UI", size=10, bold=True, color=SLATE_DARK)
    font_tbl_hdr = Font(name="Segoe UI", size=9.5, bold=True, color=SLATE_DARK)
    
    font_cell = Font(name="Segoe UI", size=9, bold=False, color=TEXT_DARK)
    font_cell_bold = Font(name="Segoe UI", size=9, bold=True, color=TEXT_DARK)
    font_cell_amount = Font(name="Segoe UI", size=9.5, bold=True, color=SLATE_DARK)
    font_subtotal = Font(name="Segoe UI", size=9.5, bold=True, color=SLATE_DARK)
    font_grand_total = Font(name="Segoe UI", size=11, bold=True, color=SLATE_DARK)
    
    font_terms_hdr = Font(name="Segoe UI", size=9.5, bold=True, color=SLATE_DARK)
    font_terms_text = Font(name="Segoe UI", size=8.5, bold=False, color="222222")
    font_footer = Font(name="Segoe UI", size=9.5, bold=True, color=SLATE_DARK)

    # ==========================================
    # --- FILLS ---
    # ==========================================
    fill_gold_hdr = PatternFill(start_color=GOLD_PRIMARY, end_color=GOLD_PRIMARY, fill_type="solid")
    fill_tbl_hdr = PatternFill(start_color=GOLD_ACCENT, end_color=GOLD_ACCENT, fill_type="solid")
    fill_subtotal = PatternFill(start_color=GOLD_TOTAL, end_color=GOLD_TOTAL, fill_type="solid")
    fill_grand_total = PatternFill(start_color=GOLD_PRIMARY, end_color=GOLD_PRIMARY, fill_type="solid")
    fill_zebra_light = PatternFill(start_color=GOLD_LIGHT, end_color=GOLD_LIGHT, fill_type="solid")
    fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    fill_dark_bar = PatternFill(start_color=SLATE_DARK, end_color=SLATE_DARK, fill_type="solid")
    fill_meta_lbl = PatternFill(start_color="EADCB5", end_color="EADCB5", fill_type="solid")

    # ==========================================
    # --- ALIGNMENTS ---
    # ==========================================
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", indent=1, wrap_text=True)
    align_left_top = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)

    # ==========================================
    # --- COLUMN WIDTHS (7 Columns: A to G) ---
    # ==========================================
    column_widths = {
        'A': 8,     # Sl No
        'B': 22,    # Item Name / Particulars
        'C': 34,    # Sub Item Name / Item Details
        'D': 14,    # Total SQFT
        'E': 16,    # Rate / Unit Price
        'F': 18,    # Amount (₹)
        'G': 32     # Remarks
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    def apply_range_style(start_cell, end_cell, font=None, fill=None, border=cell_border, alignment=None):
        start_col, start_row = openpyxl.utils.coordinate_to_tuple(start_cell)
        end_col, end_row = openpyxl.utils.coordinate_to_tuple(end_cell)
        for row_c in range(start_row, end_row + 1):
            for col_c in range(start_col, end_col + 1):
                cell = ws.cell(row=row_c, column=col_c)
                if font is not None: cell.font = font
                if fill is not None: cell.fill = fill
                if border is not None: cell.border = border
                if alignment is not None: cell.alignment = alignment

    r = 1

    # --- TOP HEADER BAR ---
    ws.row_dimensions[r].height = 26
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "QUOTATION"
    apply_range_style(f"A{r}", f"G{r}", font=font_main_title, fill=fill_dark_bar, alignment=align_center)
    r += 1

    # --- COMPANY INFO & BRAND LOGO ---
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r+1].height = 16
    ws.row_dimensions[r+2].height = 16
    ws.row_dimensions[r+3].height = 16
    ws.row_dimensions[r+4].height = 18

    # Left: Company Info
    ws[f"A{r}"] = "DECOR8 INDIA"
    ws[f"A{r}"].font = font_company
    ws[f"A{r}"].alignment = align_left

    ws[f"A{r+1}"] = "#14, Sy NO 36/1 Vasanth Vallabhnagar, Vasanthapura"
    ws[f"A{r+1}"].font = font_sub_info
    ws[f"A{r+1}"].alignment = align_left

    ws[f"A{r+2}"] = "Uttarahalli Hobli ,Bengaluru - 560061"
    ws[f"A{r+2}"].font = font_sub_info
    ws[f"A{r+2}"].alignment = align_left

    ws[f"A{r+3}"] = "8884131414 ,9380523743"
    ws[f"A{r+3}"].font = font_sub_info
    ws[f"A{r+3}"].alignment = align_left

    ws[f"A{r+4}"] = "GSTIN: 29CHPPB0944C2ZD"
    ws[f"A{r+4}"].font = font_gstin
    ws[f"A{r+4}"].alignment = align_left

    # Right: Brand Identity Block
    ws.merge_cells(f"E{r}:G{r+1}")
    ws[f"E{r}"] = "DECOR8 INDIA"
    ws[f"E{r}"].font = font_brand_rt
    ws[f"E{r}"].alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(f"E{r+2}:G{r+3}")
    ws[f"E{r+2}"] = "Affordable Luxury Interiors\nwww.decor8india.com"
    ws[f"E{r+2}"].font = font_tagline_rt
    ws[f"E{r+2}"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    r += 5

    # Decorative Divider Line
    ws.row_dimensions[r].height = 3
    apply_range_style(f"A{r}", f"G{r}", fill=fill_gold_hdr, border=None)
    r += 1

    # ==========================================
    # --- CLIENT & QUOTATION METADATA BLOCK ---
    # ==========================================
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = "QUOTATION For - MR.GOVINDA- Kolte Patil D1304"
    apply_range_style(f"A{r}", f"D{r}", font=font_banner, fill=fill_gold_hdr, alignment=align_left)

    ws.merge_cells(f"E{r}:G{r}")
    ws[f"E{r}"] = "Company Representative: Mr.Satish"
    apply_range_style(f"E{r}", f"G{r}", font=font_banner, fill=fill_gold_hdr, alignment=align_left)
    r += 1

    # Meta Rows
    meta_data = [
        ("Quotation No:", "D8202607102", "Quotation Date:", "06-AUG-2026"),
        ("Contact No:", "+91******93754", "Address:", "Bengaluru"),
        ("Email ID :", "-", "", "")
    ]

    for lbl1, val1, lbl2, val2 in meta_data:
        ws.row_dimensions[r].height = 20
        # Left Label
        ws[f"A{r}"] = lbl1
        ws[f"A{r}"].font = font_cell_bold
        ws[f"A{r}"].fill = fill_meta_lbl
        ws[f"A{r}"].border = cell_border
        ws[f"A{r}"].alignment = align_left

        # Left Value
        ws.merge_cells(f"B{r}:D{r}")
        ws[f"B{r}"] = val1
        apply_range_style(f"B{r}", f"D{r}", font=font_cell, fill=fill_white, border=cell_border, alignment=align_left)

        # Right Label
        ws[f"E{r}"] = lbl2
        ws[f"E{r}"].font = font_cell_bold
        ws[f"E{r}"].fill = fill_meta_lbl if lbl2 else fill_white
        ws[f"E{r}"].border = cell_border
        ws[f"E{r}"].alignment = align_left

        # Right Value
        ws.merge_cells(f"F{r}:G{r}")
        ws[f"F{r}"] = val2
        apply_range_style(f"F{r}", f"G{r}", font=font_cell, fill=fill_white, border=cell_border, alignment=align_left)

        r += 1

    # Spacing
    ws.row_dimensions[r].height = 6
    r += 1

    # =========================================================================
    # --- SECTION A: Wood work & Modular Finish CENTURY PLY ---
    # =========================================================================
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "Section A: Wood work & Modular Finish CENTURY PLY"
    apply_range_style(f"A{r}", f"G{r}", font=font_sec_hdr, fill=fill_gold_hdr, alignment=align_left)
    r += 1

    # Section A Table Header
    ws.row_dimensions[r].height = 22
    headers_a = ["Sl No", "Item Name", "Sub Item Name", "Total SQFT", "Rate", "Amount", "Remarks"]
    for c_idx, h_text in enumerate(headers_a, start=1):
        c_letter = get_column_letter(c_idx)
        ws[f"{c_letter}{r}"] = h_text
        ws[f"{c_letter}{r}"].font = font_tbl_hdr
        ws[f"{c_letter}{r}"].fill = fill_tbl_hdr
        ws[f"{c_letter}{r}"].border = cell_border
        ws[f"{c_letter}{r}"].alignment = align_center if c_idx in [1, 4, 5, 6] else align_left
    r += 1

    sec_a_start_row = r

    items_a = [
        (1, "Utility storage", "Profile shutters", 26, 1100, "For tv unit and kitchen"),
        ("", "", "Loft", 38, 950, "Frame with shutters"),
        ("", "", "Tall Unit", 18, 1550, "-"),
        ("", "", "Wicker Basket", "-", 12000, "2 Baskets"),
        (2, "KBR", "Swing Wardrobe", 60, 1550, "Century MR Ply carcase & Laminate Finish"),
        ("", "", "Loft", 36, 950, "Frame with shutters"),
    ]

    for sl, item, subitem, sqft, rate, rem in items_a:
        ws.row_dimensions[r].height = 20
        ws[f"A{r}"] = sl
        ws[f"B{r}"] = item
        ws[f"C{r}"] = subitem
        ws[f"D{r}"] = sqft
        ws[f"E{r}"] = rate
        if isinstance(sqft, (int, float)):
            ws[f"F{r}"] = f"=D{r}*E{r}"
        else:
            ws[f"F{r}"] = rate
        ws[f"G{r}"] = rem
        r += 1

    sec_a_end_row = r - 1

    for row_idx in range(sec_a_start_row, sec_a_end_row + 1):
        ws[f"A{row_idx}"].alignment = align_center
        ws[f"B{row_idx}"].alignment = align_center
        ws[f"C{row_idx}"].alignment = align_left
        ws[f"D{row_idx}"].alignment = align_center
        ws[f"E{row_idx}"].alignment = align_right
        ws[f"F{row_idx}"].alignment = align_right
        ws[f"G{row_idx}"].alignment = align_left

        ws[f"A{row_idx}"].font = font_cell
        ws[f"B{row_idx}"].font = font_cell_bold
        ws[f"C{row_idx}"].font = font_cell
        ws[f"D{row_idx}"].font = font_cell
        ws[f"E{row_idx}"].font = font_cell
        ws[f"F{row_idx}"].font = font_cell_amount
        ws[f"G{row_idx}"].font = font_cell

        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, 8):
            col_letter = get_column_letter(col_idx)
            ws[f"{col_letter}{row_idx}"].border = cell_border
            ws[f"{col_letter}{row_idx}"].fill = fill_zebra_light if is_even else fill_white

        if isinstance(ws[f"D{row_idx}"].value, (int, float)):
            ws[f"D{row_idx}"].number_format = "#,##0"
        if isinstance(ws[f"E{row_idx}"].value, (int, float)):
            ws[f"E{row_idx}"].number_format = "₹ #,##0.00"
        ws[f"F{row_idx}"].number_format = "₹ #,##0.00"

    ws.merge_cells(f"A{sec_a_start_row}:A{sec_a_start_row+3}")
    ws.merge_cells(f"B{sec_a_start_row}:B{sec_a_start_row+3}")

    ws.merge_cells(f"A{sec_a_start_row+4}:A{sec_a_start_row+5}")
    ws.merge_cells(f"B{sec_a_start_row+4}:B{sec_a_start_row+5}")

    # Sub Total Section A (Row 1)
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "Sub Total excluding GST"
    ws[f"F{r}"] = f"=SUM(F{sec_a_start_row}:F{sec_a_end_row})"
    ws[f"F{r}"].number_format = "₹ #,##0.00"
    apply_range_style(f"A{r}", f"E{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
    apply_range_style(f"F{r}", f"F{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
    apply_range_style(f"G{r}", f"G{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border)
    subtotal_a1_row = r
    r += 1

    # Sub Total Section A (Row 2 - Final A)
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "Sub Total excluding GST (A)"
    ws[f"F{r}"] = f"=F{subtotal_a1_row}"
    ws[f"F{r}"].number_format = "₹ #,##0.00"
    apply_range_style(f"A{r}", f"E{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
    apply_range_style(f"F{r}", f"F{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
    apply_range_style(f"G{r}", f"G{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border)
    subtotal_a_row = r
    r += 1

    # Spacing
    ws.row_dimensions[r].height = 6
    r += 1

    # =========================================================================
    # --- SECTION B: Appliances & Accessories ---
    # =========================================================================
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "Section B: Appliances & Accessories"
    apply_range_style(f"A{r}", f"G{r}", font=font_sec_hdr, fill=fill_gold_hdr, alignment=align_left)
    r += 1

    # Section B Table Header
    ws.row_dimensions[r].height = 22
    ws[f"A{r}"] = "Sl No"
    ws[f"B{r}"] = "Particulars"
    ws.merge_cells(f"C{r}:D{r}")
    ws[f"C{r}"] = "Item Details"
    ws[f"E{r}"] = "Price"
    ws[f"F{r}"] = "Amount"
    ws[f"G{r}"] = "Remarks"

    for c in range(1, 8):
        c_l = get_column_letter(c)
        ws[f"{c_l}{r}"].font = font_tbl_hdr
        ws[f"{c_l}{r}"].fill = fill_tbl_hdr
        ws[f"{c_l}{r}"].border = cell_border
        ws[f"{c_l}{r}"].alignment = align_center if c in [1, 5, 6] else align_left
    r += 1

    sec_b_start_row = r

    kitchen_items = [
        ("Hob and Chimney (Elica /Kutchina)", "-", "-", "-"),
        ("Sink", "-", "-", "-"),
        ("Inbuilt Oven and Microwave", "-", "-", "As per model selection price may vary"),
        ("Copper pipe for Gas Cylinder connection", "-", "-", "-"),
        ("1x cutlery with tandem drawer", "-", "-", "-"),
        ("1x plain ss basket drawer", "-", "-", "-"),
        ("Rolling shutter", "-", "-", "-"),
        ("1x thali basket , Magic corner , Wicker basket", "-", "-", "-")
    ]

    kitchen_start = r
    for idx, (item_det, price, amt, rem) in enumerate(kitchen_items):
        ws.row_dimensions[r].height = 20
        ws[f"A{r}"] = 1 if idx == 0 else ""
        ws[f"B{r}"] = "Kitchen\nHardware(including soft hinges )" if idx == 0 else ""
        ws.merge_cells(f"C{r}:D{r}")
        ws[f"C{r}"] = item_det
        ws[f"E{r}"] = price
        ws[f"F{r}"] = amt
        ws[f"G{r}"] = rem
        r += 1
    kitchen_end = r - 1

    ws.merge_cells(f"A{kitchen_start}:A{kitchen_end}")
    ws.merge_cells(f"B{kitchen_start}:B{kitchen_end}")

    # Wardrobes Row
    wardrobe_row = r
    ws.row_dimensions[r].height = 22
    ws[f"A{r}"] = 2
    ws[f"B{r}"] = "Wardrobes"
    ws.merge_cells(f"C{r}:D{r}")
    ws[f"C{r}"] = "Hardware-"
    ws[f"E{r}"] = 23500
    ws[f"F{r}"] = 23500
    ws[f"G{r}"] = "handles, hinges and required hardwares added"
    r += 1

    sec_b_end_row = r - 1

    for row_idx in range(sec_b_start_row, sec_b_end_row + 1):
        ws[f"A{row_idx}"].alignment = align_center
        ws[f"B{row_idx}"].alignment = align_center
        ws[f"C{row_idx}"].alignment = align_left
        ws[f"E{row_idx}"].alignment = align_center if ws[f"E{row_idx}"].value == "-" else align_right
        ws[f"F{row_idx}"].alignment = align_center if ws[f"F{row_idx}"].value == "-" else align_right
        ws[f"G{row_idx}"].alignment = align_left

        for c in range(1, 8):
            c_l = get_column_letter(c)
            ws[f"{c_l}{row_idx}"].font = font_cell
            ws[f"{c_l}{row_idx}"].border = cell_border
            ws[f"{c_l}{row_idx}"].fill = fill_zebra_light if row_idx % 2 == 0 else fill_white

        if isinstance(ws[f"E{row_idx}"].value, (int, float)):
            ws[f"E{row_idx}"].number_format = "₹ #,##0.00"
            ws[f"E{row_idx}"].font = font_cell_bold
        if isinstance(ws[f"F{row_idx}"].value, (int, float)):
            ws[f"F{row_idx}"].number_format = "₹ #,##0.00"
            ws[f"F{row_idx}"].font = font_cell_amount

    # Sub Total Section B
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "Sub Total excluding GST (B)"
    ws[f"F{r}"] = f"=SUM(F{wardrobe_row}:F{wardrobe_row})"
    ws[f"F{r}"].number_format = "₹ #,##0.00"
    apply_range_style(f"A{r}", f"E{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
    apply_range_style(f"F{r}", f"F{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
    apply_range_style(f"G{r}", f"G{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border)

    subtotal_b_row = r
    r += 1

    # Spacing
    ws.row_dimensions[r].height = 6
    r += 1

    # =========================================================================
    # --- SECTION C: Miscellaneous & Loose Furniture ---
    # =========================================================================
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "Section C: Miscellaneous & Loose Furniture"
    apply_range_style(f"A{r}", f"G{r}", font=font_sec_hdr, fill=fill_gold_hdr, alignment=align_left)
    r += 1

    # Section C Table Header
    ws.row_dimensions[r].height = 22
    ws[f"A{r}"] = "Sl No"
    ws.merge_cells(f"B{r}:D{r}")
    ws[f"B{r}"] = "Particulars"
    ws[f"E{r}"] = "Price"
    ws[f"F{r}"] = "Amount"
    ws[f"G{r}"] = "Remarks"

    for c in range(1, 8):
        c_l = get_column_letter(c)
        ws[f"{c_l}{r}"].font = font_tbl_hdr
        ws[f"{c_l}{r}"].fill = fill_tbl_hdr
        ws[f"{c_l}{r}"].border = cell_border
        ws[f"{c_l}{r}"].alignment = align_center if c in [1, 5, 6] else align_left
    r += 1

    misc_items = [
        ("King size bed Hydrolic with Headboard and side table", "-", "-", "-"),
        ("Queen size bed Hydrolic with Headboard and side table with cushion covering", "-", "-", "-"),
        ("Wallpapers", "-", "-", "-"),
        ("Painting (Royal Aspira ) 1 coat primer 2coat Asian paint Aspira", "-", "-", "-"),
        ("Quartz/Granite stone including cutting and installation", "-", "-", "-"),
        ("Electrical work ( including pop wiring switches etc )", "-", "-", "-")
    ]

    sec_c_start_row = r
    for idx, (particulars, price, amt, rem) in enumerate(misc_items, start=1):
        ws.row_dimensions[r].height = 20
        ws[f"A{r}"] = idx
        ws.merge_cells(f"B{r}:D{r}")
        ws[f"B{r}"] = particulars
        ws[f"E{r}"] = price
        ws[f"F{r}"] = 0
        ws[f"G{r}"] = rem
        r += 1
    sec_c_end_row = r - 1

    for row_idx in range(sec_c_start_row, sec_c_end_row + 1):
        ws[f"A{row_idx}"].alignment = align_center
        ws[f"B{row_idx}"].alignment = align_left
        ws[f"E{row_idx}"].alignment = align_center
        ws[f"F{row_idx}"].alignment = align_right
        ws[f"G{row_idx}"].alignment = align_left

        for c in range(1, 8):
            c_l = get_column_letter(c)
            ws[f"{c_l}{row_idx}"].font = font_cell
            ws[f"{c_l}{row_idx}"].border = cell_border
            ws[f"{c_l}{row_idx}"].fill = fill_zebra_light if row_idx % 2 == 0 else fill_white

        ws[f"F{row_idx}"].number_format = "₹ #,##0.00; - ; \"-\""

    # Sub Total Section C
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "Sub Total excluding GST (C)"
    ws[f"F{r}"] = f"=SUM(F{sec_c_start_row}:F{sec_c_end_row})"
    ws[f"F{r}"].number_format = "₹ #,##0.00"
    apply_range_style(f"A{r}", f"E{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
    apply_range_style(f"F{r}", f"F{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border, alignment=align_right)
    apply_range_style(f"G{r}", f"G{r}", font=font_subtotal, fill=fill_subtotal, border=subtotal_border)

    subtotal_c_row = r
    r += 1

    # Spacing
    ws.row_dimensions[r].height = 6
    r += 1

    # =========================================================================
    # --- TOTAL PROJECT COST SUMMARY (EXCL & INCL GST) ---
    # =========================================================================
    # Total Project Cost (A+B+C) excluding GST
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "Total Project Cost (A+B+C) excluding GST"
    ws[f"F{r}"] = f"=F{subtotal_a_row}+F{subtotal_b_row}+F{subtotal_c_row}"
    ws[f"F{r}"].number_format = "₹ #,##0.00"
    apply_range_style(f"A{r}", f"E{r}", font=font_subtotal, fill=PatternFill(start_color="E6D59E", end_color="E6D59E", fill_type="solid"), border=cell_border, alignment=align_right)
    apply_range_style(f"F{r}", f"F{r}", font=Font(name="Segoe UI", size=10, bold=True, color=SLATE_DARK), fill=PatternFill(start_color="E6D59E", end_color="E6D59E", fill_type="solid"), border=cell_border, alignment=align_right)
    apply_range_style(f"G{r}", f"G{r}", border=cell_border, fill=PatternFill(start_color="E6D59E", end_color="E6D59E", fill_type="solid"))
    total_excl_gst_row = r
    r += 1

    # CGST @ 9%
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "CGST@ 9%"
    ws[f"F{r}"] = f"=ROUND(F{total_excl_gst_row}*0.09, 0)"
    ws[f"F{r}"].number_format = "₹ #,##0.00"
    apply_range_style(f"A{r}", f"E{r}", font=font_cell_bold, fill=fill_white, border=cell_border, alignment=align_right)
    apply_range_style(f"F{r}", f"F{r}", font=font_cell_bold, fill=fill_white, border=cell_border, alignment=align_right)
    apply_range_style(f"G{r}", f"G{r}", border=cell_border, fill=fill_white)
    cgst_row = r
    r += 1

    # SGST @ 9%
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "SGST@ 9%"
    ws[f"F{r}"] = f"=ROUND(F{total_excl_gst_row}*0.09, 0)"
    ws[f"F{r}"].number_format = "₹ #,##0.00"
    apply_range_style(f"A{r}", f"E{r}", font=font_cell_bold, fill=fill_white, border=cell_border, alignment=align_right)
    apply_range_style(f"F{r}", f"F{r}", font=font_cell_bold, fill=fill_white, border=cell_border, alignment=align_right)
    apply_range_style(f"G{r}", f"G{r}", border=cell_border, fill=fill_white)
    sgst_row = r
    r += 1

    # Total Project Cost (A+B+C) including GST
    ws.row_dimensions[r].height = 26
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "Total Project Cost (A+B+C) including GST"
    ws[f"F{r}"] = f"=F{total_excl_gst_row}+F{cgst_row}+F{sgst_row}"
    ws[f"F{r}"].number_format = "₹ #,##0.00"
    apply_range_style(f"A{r}", f"E{r}", font=font_grand_total, fill=fill_grand_total, border=grand_total_border, alignment=align_right)
    apply_range_style(f"F{r}", f"F{r}", font=Font(name="Segoe UI", size=12, bold=True, color=SLATE_DARK), fill=fill_grand_total, border=grand_total_border, alignment=align_right)
    apply_range_style(f"G{r}", f"G{r}", border=grand_total_border, fill=fill_grand_total)
    grand_total_row = r
    r += 1

    # Spacing
    ws.row_dimensions[r].height = 6
    r += 1

    # =========================================================================
    # --- PAYMENT SCHEDULE ---
    # =========================================================================
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "Payment Schedule"
    apply_range_style(f"A{r}", f"G{r}", font=font_sec_hdr, fill=fill_gold_hdr, alignment=align_left)
    r += 1

    # Table Header
    ws.row_dimensions[r].height = 22
    ws[f"A{r}"] = "Sl No"
    ws.merge_cells(f"B{r}:D{r}")
    ws[f"B{r}"] = "Payment Timelines"
    ws[f"E{r}"] = "Percentage"
    ws.merge_cells(f"F{r}:G{r}")
    ws[f"F{r}"] = "Amount"

    for c in range(1, 8):
        c_l = get_column_letter(c)
        ws[f"{c_l}{r}"].font = font_tbl_hdr
        ws[f"{c_l}{r}"].fill = fill_tbl_hdr
        ws[f"{c_l}{r}"].border = cell_border
        ws[f"{c_l}{r}"].alignment = align_center if c in [1, 5, 6, 7] else align_left
    r += 1

    payment_milestones = [
        ("At the time of project/order confirmation", 0.15),
        ("At the time of starting the production/execution at factory", 0.50),
        ("After completion of production/execution at the factory or before send to the site for the installation", 0.35)
    ]

    for idx, (timeline, pct) in enumerate(payment_milestones, start=1):
        ws.row_dimensions[r].height = 24
        ws[f"A{r}"] = idx
        ws.merge_cells(f"B{r}:D{r}")
        ws[f"B{r}"] = timeline
        ws[f"E{r}"] = pct
        ws.merge_cells(f"F{r}:G{r}")
        ws[f"F{r}"] = f"=ROUND(F{grand_total_row}*E{r}, 0)"

        ws[f"A{r}"].alignment = align_center
        ws[f"B{r}"].alignment = align_left
        ws[f"E{r}"].alignment = align_center
        ws[f"F{r}"].alignment = align_right

        ws[f"A{r}"].font = font_cell
        ws[f"B{r}"].font = font_cell
        ws[f"E{r}"].font = font_cell_bold
        ws[f"F{r}"].font = font_cell_amount

        ws[f"E{r}"].number_format = "0%"
        ws[f"F{r}"].number_format = "₹ #,##0.00"

        for c in range(1, 8):
            c_l = get_column_letter(c)
            ws[f"{c_l}{r}"].border = cell_border
            ws[f"{c_l}{r}"].fill = fill_zebra_light if idx % 2 == 0 else fill_white

        r += 1

    # Spacing
    ws.row_dimensions[r].height = 6
    r += 1

    # =========================================================================
    # --- COMPANY ACCOUNT INFORMATION ---
    # =========================================================================
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "Company Account information"
    apply_range_style(f"A{r}", f"G{r}", font=font_sec_hdr, fill=fill_gold_hdr, alignment=align_left)
    r += 1

    # Account Note Row 1
    ws.row_dimensions[r].height = 18
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "Please issue A/c Payee cheque in the name of \"DECOR8 INDIA\""
    apply_range_style(f"A{r}", f"G{r}", font=Font(name="Segoe UI", size=9, bold=True, color=SLATE_DARK), fill=fill_zebra_light, border=cell_border, alignment=align_center)
    r += 1

    # Account Note Row 2
    ws.row_dimensions[r].height = 18
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "You can also make online transactions for the below mentioned account details:"
    apply_range_style(f"A{r}", f"G{r}", font=Font(name="Segoe UI", size=8.5, italic=True, color=TEXT_MUTED), fill=fill_zebra_light, border=cell_border, alignment=align_center)
    r += 1

    bank_details = [
        ("Bank Name:", "IDFC FIRST BANK"),
        ("Branch Name:", "BTM Layout"),
        ("A/C Name:", "DECOR8 INDIA"),
        ("A/C Number:", "10075641203"),
        ("A/C Type:", "Current"),
        ("IFSC Code:", "IDFB0080182")
    ]

    for lbl, val in bank_details:
        ws.row_dimensions[r].height = 20
        # Left Label (Cols A-B)
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"] = lbl
        apply_range_style(f"A{r}", f"B{r}", font=font_cell_bold, fill=fill_meta_lbl, border=cell_border, alignment=align_left)

        # Value (Cols C-E)
        ws.merge_cells(f"C{r}:E{r}")
        ws[f"C{r}"] = val
        apply_range_style(f"C{r}", f"E{r}", font=font_cell_bold if "DECOR8" in val or "IDFC" in val else font_cell, fill=fill_white, border=cell_border, alignment=align_left)

        # Empty cols F-G
        ws.merge_cells(f"F{r}:G{r}")
        apply_range_style(f"F{r}", f"G{r}", fill=fill_white, border=cell_border)

        r += 1

    # Spacing
    ws.row_dimensions[r].height = 6
    r += 1

    # =========================================================================
    # --- TERMS & CONDITIONS ---
    # =========================================================================
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "* Terms & Conditions"
    apply_range_style(f"A{r}", f"G{r}", font=font_terms_hdr, fill=fill_gold_hdr, alignment=align_center)
    r += 1

    terms_list = [
        "Price is valid for 30 days only after the 1st estimation is offered. Possibility of price revision may happen due to vendor supplier reasons.",
        "DECOR8 will cover warranty period upto 15 years for any wood work related products on manufacturing defects (Plywood bendness not included). DECOR8 will not cover warranty for any intentional wear and tear or damages from the customer side. All other products from different vendors like Hettich, KAFF, EBCO, Havells etc will be covered by respective vendors /company as per their company norms and DECOR8 will support for the same.",
        "Payment terms as per payment slab should be followed from the time of booking by the client for the smooth functioning of work and output. After any revised quote due to any addition of products/changes, if price adds up, client needs to pay the difference amount as on the slab installment during that phase. This is due to the production related requirement for procuring materials, etc. For any reductions on amount due changes, the amount will be adjusted in the final billing stage.",
        "Designs will not start until complete booking amount is credited to the company followed by the payment slabs for production & installation after progress.",
        "Any delay in terms of payment as per above payment schedule then there might be a delay in the execution for which DECOR8 wont be responsible.",
        "Price may be revised if there is any hike in the raw material announce officially by the supplier if the project is still in designing stage and is yet to enter production stage. No revision of price for projects which are under production.",
        "Once an order is confirmed and customer wants to cancel the order then NO amount will be refunded.",
        "No hidden charges for 3D designs (only 03 corrections allowed)",
        "Any corrections beyond 03 times, INR 3000 per each rendering will be charged to the client.",
        "Once the designs are approved and execution has started, no changes will be accepted in terms of structure or designs."
    ]

    for idx, term in enumerate(terms_list, start=1):
        line_count = (len(term) // 110) + 1
        ws.row_dimensions[r].height = max(20, line_count * 16)
        ws.merge_cells(f"A{r}:G{r}")
        ws[f"A{r}"] = f"* {term}"
        apply_range_style(f"A{r}", f"G{r}", font=font_terms_text, fill=fill_white if idx % 2 != 0 else fill_zebra_light, border=cell_border, alignment=align_left_top)
        r += 1

    # Spacing
    ws.row_dimensions[r].height = 6
    r += 1

    # =========================================================================
    # --- FOOTER BANNER ---
    # =========================================================================
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "www.decor8india.com"
    apply_range_style(f"A{r}", f"G{r}", font=font_footer, fill=fill_gold_hdr, alignment=align_center)
    r += 1

    # Safe save
    try:
        wb.save(output_path)
        print(f"Saved: {output_path}")
    except PermissionError:
        alt_path = output_path.replace(".xlsx", "_V2.xlsx")
        wb.save(alt_path)
        print(f"Original locked. Saved to alternate: {alt_path}")

if __name__ == "__main__":
    build_perfect_decor8_quotation("DECOR8_INDIA_QUOTATION_MR_GOVINDA_D1304.xlsx")
    build_perfect_decor8_quotation("public/DECOR8_INDIA_QUOTATION_MR_GOVINDA_D1304.xlsx")
